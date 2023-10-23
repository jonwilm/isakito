from ast import While
import os
from openpyxl import load_workbook
from apps.productos.models import Brand, Product, ImagesProduct, ImportProducts
from decimal import Decimal


def run(*args):
    files = ImportProducts.objects.filter(status='1')
    if len(files) == 0:
        return False
    for file in files:
        module_dir = os.path.dirname(__file__)
        file_path = os.path.join(module_dir, '..' + file.document.url)

        wb = load_workbook(
            filename=file_path,
            read_only=True)
        ws = wb[wb.sheetnames[0]]

        try:
            for row in ws.rows:
                if str(row[0].value).lower() in ['none', '-', '']:
                    break

                if str(row[0].value).lower() != 'sku':
                    brand, create_brand = Brand.objects.update_or_create(
                        name=str(row[7].value),
                    )
                    product, create_product = Product.objects.update_or_create(
                        sku=str(row[0].value),
                        defaults={
                            'title': str(row[1].value),
                            'description': str(row[2].value),
                            'price': str(row[3].value),
                            'ean': str(row[4].value),
                            'category': str(row[6].value),
                            'brand': brand,
                            'stock': str(row[8].value),
                            'supplier': str(row[9].value),
                            'height': str(row[10].value),
                            'width': str(row[11].value),
                            'long': str(row[12].value),
                            'weight': str(row[13].value),
                        }
                    )
                    images = str(row[5].value).split(",")
                    # print(str(row[0].value))
                    # print(images)
                    for image in images:
                        image_product, create_image_product = ImagesProduct.objects.update_or_create(
                            image=image,
                            product=product,
                            defaults={
                            }
                        )
                        # print('Imagen' + image)
            print("Importacion Completa")
            file.status = '2'
            file.save()
        except Exception as exc:
            print("Importacion Fallida: ", exc)
            # file.status = '3'
            # file.save()