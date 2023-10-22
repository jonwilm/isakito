import os
from openpyxl import load_workbook
from apps.productos.models import Marca, Producto, ImagenesProducto, ImportarProductos
from decimal import Decimal


def get_row_value(row, i):
    try:
        if str(row[i].value) == 'None':
            return ''
        return str(row[i].value)
    except:
        return "-"


def run(*args):
    files = ImportarProductos.objects.filter(status='1')
    if len(files) == 0:
        return False
    for file in files:
        module_dir = os.path.dirname(__file__)
        file_path = os.path.join(module_dir, '..' + file.documento.url)
        print(module_dir)
        print(file_path)
        print(file.documento.url)

        wb = load_workbook(
            filename=file_path,
            read_only=True)
        ws = wb[wb.sheetnames[0]]

        try:
            for row in ws.rows:
                if get_row_value(row, 0).lower() != 'sku':
                    marca, create_marca = Marca.objects.update_or_create(
                        nombre=get_row_value(row, 7),
                    )
                    producto, create_producto = Producto.objects.update_or_create(
                        sku=get_row_value(row, 0),
                        defaults={
                            'titulo': get_row_value(row, 1),
                            'descripcion': get_row_value(row, 2),
                            'precio': get_row_value(row, 3),
                            'ean': get_row_value(row, 4),
                            'categoria': get_row_value(row, 6),
                            'marca': marca,
                            'stock': get_row_value(row, 8),
                            'proveedor': get_row_value(row, 9),
                            'alto': get_row_value(row, 10),
                            'ancho': get_row_value(row, 11),
                            'largo': get_row_value(row, 12),
                            'peso': get_row_value(row, 13),
                        }
                    )
                    images = get_row_value(row, 5).split(",")
                    for image in images:
                        images_producto, create_images_producto = ImagenesProducto.objects.update_or_create(
                            imagen=image,
                            producto=producto,
                            defaults={
                            }
                        )
                        print(image)
            print("Importacion Completa")
            file.status = '2'
            file.save()
        except Exception as exc:
            print("Importacion Fallida: ", exc)
            # file.status = '3'
            # file.save()