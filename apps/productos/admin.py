import os
from openpyxl import load_workbook
from django.contrib import admin
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from django.db import models

from io import BytesIO

from .models import Brand, SliderBrand, Product, ImagesProduct, ImportProducts


class SliderBrandResource(resources.ModelResource):
    class Meta:
        model: SliderBrand


class SliderBrandAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    list_display = ('brand', 'show_img', 'title', 'order', 'active')
    exclude = ('model_state',)
    list_filter = ('brand',)
    ordering = ('-active',)


class SliderBrandInlineAdmin(admin.TabularInline):
    model = SliderBrand
    exclude = ('model_state',)
    extra = 0


class BrandResource(resources.ModelResource):
    class Meta:
        model: Brand


class BrandAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    inlines = (SliderBrandInlineAdmin,)
    list_display = ('name', 'show_img', 'active')
    exclude = ('model_state', 'slug')
    list_filter = ('name',)
    ordering = ('-active',)


class ImagenesProductInlineAdmin(admin.TabularInline):
    model = ImagesProduct
    exclude = ('model_state',)
    extra = 0


class ProductResource(resources.ModelResource):
    class Meta:
        model: Product


class ProductAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    inlines = (ImagenesProductInlineAdmin,)
    list_display = ('sku', 'title', 'brand', 'active')
    exclude = ('model_state', 'slug')
    list_filter = ('brand',)
    ordering = ('-active',)


class ImagenesProductAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    list_display = ('product', 'image', 'active')
    exclude = ('model_state',)
    list_filter = ('product',)
    ordering = ('-active',)


@admin.action(description='Importar Excels seleccionados')
def import_products(modeladmin, request, queryset):
    files = queryset
    for file in files:
        wb = load_workbook(
            filename=file.document,
            read_only=True)
        ws = wb[wb.sheetnames[0]]

        try:
            for row in ws.rows:
                if str(row[0].value).lower() in ['none', '-', '']:
                    break

                if str(row[0].value).lower() != 'sku':
                    images = str(row[5].value).split(",")
                    imageProduct = images[0]
                    brand, create_brand = Brand.objects.update_or_create(
                        name=str(row[7].value),
                    )
                    product, create_product = Product.objects.update_or_create(
                        sku=str(row[0].value),
                        defaults={
                            'title': str(row[1].value),
                            'image': imageProduct,
                            'description': str(row[2].value),
                            'price': str(row[3].value),
                            'ean': str(row[4].value),
                            'category': str(row[6].value),
                            'brand': brand,
                            'stock': str(row[8].value),
                            'supplier': str(row[9].value),
                            'height': str(row[10].value),
                            'width': str(row[11].value),
                            'long': str(row[12].value),
                            'weight': str(row[13].value),
                        }
                    )
                    for image in images:
                        image_product, create_image_product = ImagesProduct.objects.update_or_create(
                            image=image,
                            product=product,
                            defaults={
                            }
                        )
            print("Importacion Completa")
            file.status = '2'
            file.save()
        except Exception as exc:
            print("Importacion Fallida: ", exc)
            file.status = '3'
            file.save()

class ImportarProductsAdmin(admin.ModelAdmin):
    list_display = ('date_created', 'document', 'status')
    exclude = ('model_state', 'status')
    ordering = ('status', '-date_created')
    actions = [import_products,]


admin.site.register(Brand, BrandAdmin)
# admin.site.register(SliderBrand, SliderBrandAdmin)
admin.site.register(Product, ProductAdmin)
# admin.site.register(ImagesProduct, ImagenesProductAdmin)
admin.site.register(ImportProducts, ImportarProductsAdmin)
